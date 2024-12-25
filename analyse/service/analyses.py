import glob
import logging
import os
import queue
import re
from collections import defaultdict
from datetime import datetime
from io import BytesIO
import pytesseract
from PIL import Image
from django.utils import timezone
from pyxirr import xirr

from _decimal import Decimal
from matplotlib import ticker
from openpyxl.reader.excel import load_workbook
from scipy.optimize import newton

import data.transfer
from analyse.bo import OperateItem, CostOfLevelInfo
from analyse.models import OperateRecord, PurchaseInRecord, PurchaseOutRecord, MatchRelations
import numpy as np
import matplotlib.pyplot as plt

logger = logging.getLogger("django")

class WealthException(Exception):
    pass


# 从[买入]操作中找到与卖出记录匹配的买入记录
# 1. 时间最近 且 盈利的 记录,如果不存在,跳到2
# 2. 时间最近 的 记录,如果不存在,跳到3
# 3. 抛出异常
def getMatchForPurchaseOut(operateRecordHead, purchaseInOperateItemWaitForMatch):
    # 将买入记录按交易时间降序排序
    purchaseInOperateItemWaitForMatchSorted = sorted(purchaseInOperateItemWaitForMatch, key=lambda temp: temp.operateTime, reverse=True)

    # 找到第一条盈利的买入记录
    matchedOperateItem = next((obj for obj in purchaseInOperateItemWaitForMatchSorted if obj.unitPrice < operateRecordHead.unitPrice), None)
    if matchedOperateItem is not None:
        return matchedOperateItem

    # 找到第一条买入记录
    return purchaseInOperateItemWaitForMatchSorted[0]


# 将亏损金额平摊到未匹配的PurchaseIn集合中
def amortizeToPurchaseInOperateItem(loss, purchaseInOperateItemWaitForMatch):
    # 设置每次分摊的亏损金额,单位:元
    amortizeAmount = 1

    while loss > 0:
        # 获取本轮实际要分摊的亏损
        trueLoss = amortizeAmount if loss > amortizeAmount else loss

        # 获取筹码水平最低的待匹配记录
        operateItem = min(purchaseInOperateItemWaitForMatch, key=lambda OperateItem: OperateItem.unitPrice)

        # 新成本价= 亏损金额/100 + 原成本价
        operateItem.unitPrice = Decimal((str)(trueLoss / 100)) + operateItem.unitPrice

        # 减掉已经分摊的亏损金额
        loss -= trueLoss


def analyseSpecificBond(bondCode):
    # 从数据库中查询该证劵交易编码对应的交易记录 按交易时间升序排序
    operateRecordList = OperateRecord.objects.filter(bond_code=bondCode).order_by('bargain_time')

    # 按序取出交易记录,按100份为单位拆分,放进队列中
    totalFee = 0
    queueFor100Unit = queue.Queue(-1)
    for operateRecord in operateRecordList:
        splitCount = operateRecord.bargain_number // 100
        for _ in range(splitCount):
            queueFor100Unit.put(OperateItem(
                operateKind=operateRecord.operate_kind,
                operateTime=operateRecord.bargain_time,
                unitPrice=operateRecord.bargain_unit_price,
                share=100,
                fee=0
            ))
        # 统计总手续费
        totalFee += operateRecord.fee

    purchaseInOperateItemWaitForMatch = []
    profit = 0
    loss = 0
    profit_in_cost = 0

    # 用于计算已确认收益XIRR
    confirm_transaction_time = []
    confirm_cash_flow = []
    # 取出队头交易记录
    while not queueFor100Unit.empty():
        operateRecordHead = queueFor100Unit.get()
        # a. 如果操作类型为[买入]
        if operateRecordHead.operateKind == 'in':
            # 放进待匹配的list中
            purchaseInOperateItemWaitForMatch.append(operateRecordHead)
        # b. 如果操作类型为[卖出]
        elif operateRecordHead.operateKind == 'out':
            # 从[买入]操作中找到匹配的买入记录
            purchaseInOperateItemMatched = getMatchForPurchaseOut(operateRecordHead = operateRecordHead,
                                                                  purchaseInOperateItemWaitForMatch=purchaseInOperateItemWaitForMatch)
            # 将匹配的[买入]操作从待匹配记录中移除
            purchaseInOperateItemWaitForMatch.remove(purchaseInOperateItemMatched)
            # 记录总亏损/总盈亏
            inAmount = purchaseInOperateItemMatched.unitPrice
            outAmount = operateRecordHead.unitPrice
            # 买入>卖出 亏损
            if inAmount > outAmount:
                loss += (inAmount - outAmount)*100
                # 买入价降低至卖出价,保证此次交易不亏损
                profit_in_cost += outAmount * 100

                # 计算xirr使用
                # 买入流水
                confirm_cash_flow.append(-outAmount*100)
                confirm_transaction_time.append(purchaseInOperateItemMatched.operateTime)
                # 卖出流水
                confirm_cash_flow.append(outAmount*100)
                confirm_transaction_time.append(operateRecordHead.operateTime)
            # 买入<=卖出 盈利
            else:
                profit += (outAmount - inAmount)*100
                profit_in_cost += inAmount * 100

                # 计算xirr使用
                # 买入流水
                confirm_cash_flow.append(-inAmount*100)
                confirm_transaction_time.append(purchaseInOperateItemMatched.operateTime)
                # 卖出流水
                confirm_cash_flow.append(outAmount*100)
                confirm_transaction_time.append(operateRecordHead.operateTime)
        else:
            raise WealthException("操作类型无法识别", operateRecordHead.operateKind)

    # 计算XIRR
    try:
        xirrv = xirr(zip(confirm_transaction_time, confirm_cash_flow))
    except Exception as e:
        print("计算XIRR时出错:", e)

    print("xirr:", xirrv)

    # 将亏损金额平摊到未匹配的PurchaseIn集合中
    # 如果所有买入记录均已匹配,则不需要均摊
    if purchaseInOperateItemWaitForMatch:
        amortizeToPurchaseInOperateItem(loss, purchaseInOperateItemWaitForMatch)

    # 按持有成本进行汇总
    costOfLevelList = []
    if purchaseInOperateItemWaitForMatch:
        # 将purchaseInOperateItemWaitForMatch这个List中unitPrice相等的元素的share进行累计
        for tempOperateRecord in purchaseInOperateItemWaitForMatch:
            tempUnitPrice = tempOperateRecord.unitPrice
            tempShare = tempOperateRecord.share
            # 判断这个筹码水平在costOfLevelList中是否存在
            existForLevel = False
            for tempCostOfLevel in costOfLevelList:
                # 如果已经存在,则取出修改份额
                if tempCostOfLevel.unitPrice == tempUnitPrice:
                    existForLevel = True
                    tempCostOfLevel.share += tempShare
            # 如果不存在,则新增一个筹码水平对象
            if not existForLevel:
                costOfLevelList.append(CostOfLevelInfo(
                    unitPrice=tempUnitPrice,
                    share=tempShare))
        # 将costOfLevelList按unitPrice升序排序,从低到高
        costOfLevelList = sorted(costOfLevelList, key=lambda CostOfLevelInfo: CostOfLevelInfo.unitPrice)
        for temp in costOfLevelList:
            print("unitPrice:", temp.unitPrice, "share:", temp.share)
        # print(bondCode + "筹码水平为:" +costOfLevelList)
    confirm_profit_rate = profit/profit_in_cost
    percentage = "{:.2%}".format(confirm_profit_rate)
    print("手续费:", totalFee, "盈利:", profit, "亏损:", loss, "已确认收益率", percentage)

    # 获取最新的收盘价
    presentPrice = data.transfer.getAllDayDataOfSpecificETF(bondCode).iloc[-1, 4]

    # 筹码水平可视化
    # x轴:份额 y轴:价格
    unitPrice = [float(costOfLevel.unitPrice) for costOfLevel in costOfLevelList]
    share = [float(costOfLevel.share) for costOfLevel in costOfLevelList]
    # plt.gca().yaxis.set_major_locator(ticker.MultipleLocator(base=15))
    plt.barh(unitPrice, share, height=0.001, color='blue')
    plt.ylabel('price')
    plt.xlabel('share')
    plt.title('Shares Distribution of Different Prices')

    plt.axhline(y=presentPrice, color='red', linestyle='--', label='present price')

    plt.show()

    return costOfLevelList


def addOperateRecord(file_obj):
    # 读取文件，写入数据库
    rb = load_workbook(filename=BytesIO(file_obj.read()))
    # 这个属性在处理只有一个工作表的Excel文件或者你只对最后被操作的工作表感兴趣时特别有用。它允许你不必知道工作表的具体名称或索引就能直接访问它
    sheet = rb.active

    importTotalCount = 0
    for index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        try:
            if index == 1:
                continue
            # 解析每列数据并生成新对象
            newOperateRecord = OperateRecord()
            # a. 成交时间
            newOperateRecord.bargain_time = datetime.strptime(str(row[0]) + " " + row[1].strftime("%H:%M:%S"), "%Y%m%d %H:%M:%S")
            # b. 证劵代码
            newOperateRecord.bond_code = str(row[2])
            # c. 证劵名称
            newOperateRecord.bond_name = row[3]
            # d. 操作类型
            if row[4] == '卖出':
                newOperateRecord.operate_kind = 'out'
            elif row[4] == '买入':
                newOperateRecord.operate_kind = 'in'
            else:
                raise WealthException("无法识别的操作类型", row[4])
            # e. 成交数量
            newOperateRecord.bargain_number = row[5]
            # f. 成交均价
            newOperateRecord.bargain_unit_price = row[6]
            # g. 成交金额
            newOperateRecord.bargain_amount = row[7]
            # h. 合同编号
            newOperateRecord.contract_no = str(row[8])
            # i. 成交编号
            newOperateRecord.bargain_no = str(row[9])
            # j. 手续费
            newOperateRecord.fee = row[10]
            # k. 印花税
            newOperateRecord.stamp_duty = row[11]
            # l. 其它杂费
            newOperateRecord.other_fee = row[12]
            # m. 发生金额
            newOperateRecord.happen_amount = row[13]
            # n. 资金余额
            newOperateRecord.left_amount = row[14]
            # o. 交易市场
            newOperateRecord.trade_market = row[15]
            # p. 股东账户
            newOperateRecord.share_holder_account = str(row[16])
            # q. 交收日期
            newOperateRecord.delivery_date = datetime.strptime(str(row[17]), "%Y%m%d")
            # r. 证劵中文全称
            newOperateRecord.bond_full_name = row[18]

            # 其它字段
            newOperateRecord.ctime = datetime.now()
            newOperateRecord.mtime = datetime.now()
            # 保存到数据库中
            newOperateRecord.save()

            importTotalCount += 1
            print(index, row)
        except Exception as e:
            logger.exception("导入交易记录异常: %d %s", index, e)
            break


    return {'code': 200, 'data': {}, 'msg': '导入' + str(importTotalCount) + '条成功'}


def analyseCostChangeOfSpecificBond(bondCode):
    # 从数据库中查询该证劵交易编码对应的交易记录 按交易时间升序排序
    operateRecordList = OperateRecord.objects.filter(bond_code=bondCode).order_by('bargain_time')

    # 计算每日持仓成本
    daily_cost = calculate_daily_cost(operateRecordList)
    holding_cost = get_daily_holding_cost(daily_cost)

    # 准备绘图数据
    dates = [item[0] for item in holding_cost]
    costs = [item[1] for item in holding_cost]

    # 设置中文字体
    plt.rcParams['font.sans-serif'] = ['SimHei']  # 使用黑体
    plt.rcParams['axes.unicode_minus'] = False  # 解决负号显示问题

    bond_code = operateRecordList[0].bond_code
    bond_name = operateRecordList[0].bond_name

    # 确定持仓成本的日期范围
    min_date = min(dates)
    max_date = max(dates)

    # 获取该证劵股价
    originalData = data.transfer.getAllDayDataOfSpecificETF(bond_code)
    close_prices = originalData.iloc[:, 4]
    fund_date = originalData.iloc[:, 2]
    # 将字符串日期转换为 datetime.date 类型
    fund_date = [datetime.strptime(date, "%Y-%m-%d").date() for date in fund_date]

    # 过滤实际股价数据
    filtered_close_prices = []
    filtered_fund_date = []

    for date, price in zip(fund_date, close_prices):
        if min_date <= date <= max_date:
            filtered_fund_date.append(date)
            filtered_close_prices.append(price)

    # 绘制折线图
    plt.figure(figsize=(10, 5))
    plt.plot(dates, costs, marker='o', label='持仓成本')
    plt.plot(filtered_fund_date, filtered_close_prices, marker='s', label='实时价格')
    plt.xlabel('日期')
    plt.ylabel('持仓成本')
    plt.title(f'持仓成本随时间的变化 - {bond_name} ({bond_code})')
    plt.grid(True)
    plt.xticks(rotation=45)
    plt.legend()
    plt.tight_layout()
    plt.show()

# 获取每日持仓成本
def get_daily_holding_cost(daily_cost):
    dates = sorted(daily_cost.keys())
    holding_cost = []
    total_cost = 0
    total_quantity = 0

    for date in dates:
        total_cost += daily_cost[date]['total_cost']
        total_quantity += daily_cost[date]['total_quantity']
        if total_quantity != 0:
            holding_cost.append((date, total_cost / total_quantity))
        else:
            holding_cost.append((date, 0))

    return holding_cost


# 计算每日持仓成本
def calculate_daily_cost(records):
    daily_cost = defaultdict(lambda: {'total_cost': 0, 'total_quantity': 0})

    for record in records:
        date = record.bargain_time.date()
        if record.operate_kind == 'in':
            daily_cost[date]['total_cost'] += record.bargain_amount
            daily_cost[date]['total_quantity'] += record.bargain_number
        elif record.operate_kind == 'out':
            daily_cost[date]['total_cost'] -= record.bargain_amount
            daily_cost[date]['total_quantity'] -= record.bargain_number

    return daily_cost

def xirr_calculate(bondCode):
    # 从数据库中查询该证劵交易编码对应的交易记录 按交易时间升序排序
    operateRecordList = OperateRecord.objects.filter(bond_code=bondCode).order_by('bargain_time')

    transaction_time = []
    directions = []
    amounts = []

    for record in operateRecordList:
        # 将日期转换为offset-naive
        naive_bargain_time = record.bargain_time.replace(tzinfo=None)
        transaction_time.append(naive_bargain_time)
        directions.append(record.operate_kind)
        amounts.append(record.bargain_amount)

    cash_flows = [float(-amount) if direction == 'in' else float(amount) for direction, amount in zip(directions, amounts)]

    # 获取最新的收盘价
    presentPrice = float(data.transfer.getAllDayDataOfSpecificETF(bondCode).iloc[-1, 4])  # 将Decimal转换为float

    # 计算当前持有股票的市值
    current_holdings = sum(record.bargain_number for record in operateRecordList if record.operate_kind == 'in') - \
                       sum(record.bargain_number for record in operateRecordList if record.operate_kind == 'out')
    current_value = current_holdings * presentPrice

    # 添加当前持有股票的市值
    cash_flows.append(current_value)
    transaction_time.append(datetime.strptime('2024-10-08', '%Y-%m-%d'))

    # 计算XIRR
    try:
        xirrv = xirr(zip(transaction_time, cash_flows))
    except Exception as e:
        print("计算XIRR时出错:", e)

    print("xirr:", xirrv)



def initialize_outer_left_match(bondCode):
    # 初始化买入记录的待匹配份额
    for record in PurchaseInRecord.objects.filter(bond_code=bondCode):
        record.left_match = record.purchase_in_count
        record.save()

    # 初始化卖出记录的待匹配份额
    for record in PurchaseOutRecord.objects.filter(bond_code=bondCode):
        record.left_match = record.purchase_out_count
        record.save()

    # 删除匹配记录表中的所有记录
    MatchRelations.objects.filter(bond_code=bondCode).delete()

def outer_analyse(bondCode):
    # 初始化待匹配记录
    initialize_outer_left_match(bondCode=bondCode)

    # 取出未匹配完的卖出记录，按时间由近到远排序
    unmatched_sell_records = PurchaseOutRecord.objects.filter(
        left_match__gt=0,
        bond_code=bondCode
    ).order_by('-purchase_out_time')

    total_loss = 0
    total_profit = 0

    for sell_record in unmatched_sell_records:
        sell_unmatched_quantity = sell_record.left_match
        sell_time = sell_record.purchase_out_time
        sell_price = sell_record.purchase_out_single_price

        # 取出未匹配完的买入记录，按时间由近到远排序
        unmatched_buy_records = PurchaseInRecord.objects.filter(
            left_match__gt=0,
            bond_code=bondCode
        ).order_by('-purchase_in_time')

        # 遍历卖出记录
        for buy_record in unmatched_buy_records:
            buy_unmatched_quantity = buy_record.left_match
            buy_time = buy_record.purchase_in_time
            buy_price = buy_record.purchase_in_single_price

            # 遍历买入记录
            if buy_time <= sell_time and (
                    buy_price < sell_price or not unmatched_buy_records.filter(purchase_in_time__lt=sell_time,
                                                                               purchase_in_single_price__lt=sell_price).exists()):
                # 计算匹配数量
                match_quantity = min(sell_unmatched_quantity, buy_unmatched_quantity)

                # 计算盈亏金额
                profit_or_loss = (sell_price - buy_price) * match_quantity

                if profit_or_loss > 0:
                    # 盈利
                    total_profit += profit_or_loss
                    print(f"盈利: {profit_or_loss}, 匹配份额: {match_quantity}")
                else:
                    # 亏损
                    total_loss += abs(profit_or_loss)
                    print(f"亏损: {abs(profit_or_loss)}, 匹配份额: {match_quantity}")

                # 更新匹配记录
                MatchRelations.objects.create(
                    in_uniq_id=buy_record.id,
                    out_uniq_id=sell_record.id,
                    match_count=match_quantity,
                    bond_code=bondCode
                )

                # 更新买入记录的待匹配份额
                buy_record.left_match -= match_quantity
                buy_record.save()

                # 更新卖出记录的待匹配份额
                sell_record.left_match -= match_quantity
                sell_unmatched_quantity -= match_quantity
                sell_record.save()

                # 如果卖出记录已完全匹配，跳出循环
                if sell_record.left_match == 0:
                    break

    # 将总亏损金额平均分摊到未匹配的买入记录的单价当中
    unmatched_buy_records = PurchaseInRecord.objects.filter(left_match__gt=0)
    if unmatched_buy_records.exists():
        average_loss_per_share = total_loss / sum([record.left_match for record in unmatched_buy_records])

        # 计算每个买入记录的真实单价
        real_prices = []
        for buy_record in unmatched_buy_records:
            real_price = buy_record.purchase_in_single_price + average_loss_per_share
            real_prices.append((buy_record.id, buy_record.left_match, real_price))

        # 按真实单价从小到大排序
        real_prices.sort(key=lambda x: x[2])

        # 打印买入记录的待匹配份额的真实单价，并计算相邻两个价格间的百分比差额
        for i, (id, left_match, real_price) in enumerate(real_prices):
            if i > 0:
                prev_real_price = real_prices[i - 1][2]
                percentage_diff = ((real_price - prev_real_price) / prev_real_price) * 100
                print(
                    f"买入记录ID: {id}, 待匹配份额: {left_match}, 真实单价: {real_price:.4f} ({percentage_diff:.2f}%)")
            else:
                print(f"买入记录ID: {id}, 待匹配份额: {left_match}, 真实单价: {real_price:.4f}")

    print("总亏损金额:", total_loss)
    print("总利润金额:", total_profit)

    return total_loss


def rename_file():
    # 使用 glob 模块来匹配 .jfif 文件
    jfif_files = glob.glob(os.path.join(r"/pictures/中证500场外", '**/*.jfif'), recursive=True)

    # 确保路径是绝对路径
    jfif_files = [os.path.abspath(file) for file in jfif_files]

    # 按照文件名排序，确保重命名时顺序正确
    jfif_files.sort()

    # 重命名文件
    for index, file_path in enumerate(jfif_files, start=1):
        # 获取文件所在的目录和文件名
        directory, filename = os.path.split(file_path)

        # 构建新的文件名
        new_filename = f"{index}.jfif"
        new_file_path = os.path.join(directory, new_filename)

        # 重命名文件
        os.rename(file_path, new_file_path)
        print(f"Renamed {file_path} to {new_file_path}")


def outer_ocr():
    # 设置 TESSDATA_PREFIX 环境变量
    os.environ['TESSDATA_PREFIX'] = r'F:\software\tessdata'
    # 显式指定 Tesseract 的路径
    pytesseract.pytesseract.tesseract_cmd = r'F:\software\tesseract.exe'

    # 使用 glob 模块来匹配 .jfif 文件
    jfif_files = glob.glob(os.path.join(r"F:\中证500场外", '**/*.jfif'), recursive=True)

    # 确保路径是绝对路径
    images = [os.path.abspath(file) for file in jfif_files]

    # images = [r'F:\中证500场外\15.jfif',
    #           r'F:\中证500场外\17.jfif',
    #           r'F:\中证500场外\9.jfif']  # 替换为实际图片路径

    for image in images:
        try:
            print("Processing:", image)
            text = extract_text_from_image(image)
            # 买入信息的正则表达式模式
            purchase_in_pattern = re.compile(r"""
                买\s*入\s*时\s*间\s*(?P<time>\d{4}-\d{2}-\d{2}\s\d{2}:\d{2}:\d{2})
                .*?
                确\s*认\s*份\s*额\s*(?P<shares>\d{1,3}(,\d{3})*(\.\d+)?)\s*份
                .*?
                确\s*认\s*净\s*值\s*(?P<price>\d+\.\d+)
                .*?
                手\s*续\s*费\s*(?P<fee>\d+\.\d+)\s*元
                .*?
                订\s*单\s*号\s*(?P<uniq_no>\d+)
            """, re.VERBOSE | re.DOTALL)

            # 卖出信息的正则表达式模式
            purchase_out_pattern = re.compile(r"""
                卖\s*出\s*时\s*间\s*(?P<time>\d{4}-\d{2}-\d{2}\s\d{2}:\d{2}:\d{2})
                .*?
                确\s*认\s*份\s*额\s*(?P<shares>\d+\.\d+)\s*份
                .*?
                确\s*认\s*净\s*值\s*(?P<price>\d+\.\d+)
                .*?
                手\s*续\s*费\s*.*?\s*(?P<fee>\d+\.\d+)\s*元
                .*?
                订\s*单\s*号\s*(?P<uniq_no>\d+)
            """, re.VERBOSE | re.DOTALL)

            purchase_in_match = purchase_in_pattern.search(text)
            if purchase_in_match:
                purchase_in_data = purchase_in_match.groupdict()
                purchase_in_data['purchase_in_time'] = timezone.make_aware(datetime.strptime(purchase_in_data['time'], '%Y-%m-%d %H:%M:%S'))
                # shares可能包含千分位的数字，所以需要处理
                purchase_in_data['shares'] = purchase_in_data['shares'].replace(',', '')
                purchase_in_data['purchase_in_count'] = float(purchase_in_data['shares'])
                # price可能包含千分位的数字，所以需要处理
                purchase_in_data['price'] = purchase_in_data['price'].replace(',', '')
                purchase_in_data['purchase_in_single_price'] = float(purchase_in_data['price'])
                purchase_in_data['purchase_fee'] = float(purchase_in_data['fee'])
                purchase_in_data['purchase_in_uniq_no'] = purchase_in_data['uniq_no']
                purchase_in_data['left_match'] = float(purchase_in_data['shares'])

                # 创建买入记录
                PurchaseInRecord.objects.create(
                    purchase_in_time=purchase_in_data['purchase_in_time'],
                    purchase_in_count=purchase_in_data['purchase_in_count'],
                    purchase_in_single_price=purchase_in_data['purchase_in_single_price'],
                    purchase_fee=purchase_in_data['purchase_fee'],
                    purchase_in_uniq_no=purchase_in_data['purchase_in_uniq_no'],
                    left_match=purchase_in_data['left_match'],
                    ctime=timezone.make_aware(datetime.now()),
                    mtime=timezone.make_aware(datetime.now())
                )
                print("买入插入")
                continue

            # 解析卖出信息
            purchase_out_match = purchase_out_pattern.search(text)
            if purchase_out_match:
                purchase_out_data = purchase_out_match.groupdict()
                purchase_out_data['purchase_out_time'] = timezone.make_aware(datetime.strptime(purchase_out_data['time'], '%Y-%m-%d %H:%M:%S'))
                purchase_out_data['purchase_out_count'] = float(purchase_out_data['shares'])
                purchase_out_data['purchase_out_single_price'] = float(purchase_out_data['price'])
                purchase_out_data['purchase_fee'] = float(purchase_out_data['fee'])
                purchase_out_data['purchase_out_uniq_no'] = purchase_out_data['uniq_no']
                purchase_out_data['left_match'] = float(purchase_out_data['shares'])

                # 创建卖出记录
                PurchaseOutRecord.objects.create(
                    purchase_out_time=purchase_out_data['purchase_out_time'],
                    purchase_out_count=purchase_out_data['purchase_out_count'],
                    purchase_out_single_price=purchase_out_data['purchase_out_single_price'],
                    purchase_fee=purchase_out_data['purchase_fee'],
                    purchase_out_uniq_no=purchase_out_data['purchase_out_uniq_no'],
                    left_match=purchase_out_data['left_match'],
                    ctime=timezone.make_aware(datetime.now()),
                    mtime=timezone.make_aware(datetime.now())
                )
                print("卖出插入")
                continue
        except Exception as e:
            print(f"处理图片 {image} 时发生错误: {e}")



def extract_text_from_image(image_path):
    """从图片中提取文本"""
    image = Image.open(image_path)
    text = pytesseract.image_to_string(image, lang='chi_sim')
    return text